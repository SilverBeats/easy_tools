import csv
import json
import os
import pickle
from typing import Any, Callable, Iterable, List, Union

import numpy as np
import pandas as pd
from omegaconf import DictConfig, OmegaConf
from openpyxl.reader.excel import load_workbook

from .errors import FileReadError, FileTypeError, FileWriteError
from .tools import get_file_name_and_ext

__all__ = [
    'read_json',
    'dump_json',
    'read_jsonl',
    'dump_jsonl',
    'read_txt',
    'dump_txt',
    'read_config',
    'dump_config',
    'read_pkl',
    'dump_pkl',
    'read_large_excel',
    'read_excel',
    'dump_excel',
    'read_csv',
    'dump_csv',
    'read_npyz',
    'dump_npyz',
    'FileReader',
    'FileWriter'
]


def ext_check(ext: Union[str, List[str]]):
    def decorator(func: Callable):
        def wrapper(*args, **kwargs):
            file_path = kwargs.get('file_path', None)
            if file_path is None and len(args) > 0:
                func_name = func.__name__
                file_path_index = 0 if 'read' in func_name else 1
                file_path = args[file_path_index]

            if not file_path:
                raise FileReadError(f'{file_path} is required')

            file_ext = get_file_name_and_ext(file_path, with_dot=False)[-1]

            allowed_exts = [ext] if isinstance(ext, str) else ext
            if file_ext not in allowed_exts:
                allowed_str = ', '.join(allowed_exts)
                raise FileTypeError(f'{file_path} is not a {allowed_str} file!')

            return func(*args, **kwargs)

        return wrapper

    return decorator


@ext_check(ext='json')
def read_json(file_path: str) -> Any:
    """
    Args:
        file_path: json file path
    """
    with open(file_path, 'r', encoding='utf-8') as json_file:
        return json.load(json_file)


@ext_check(ext='json')
def dump_json(data: Any, file_path: str, json_encoder=None):
    """
    Args:
        file_path: json file path
        data: data to dump
        json_encoder: for custom situation
    """
    os.makedirs(os.path.dirname(file_path) or './', exist_ok=True)
    with open(file_path, 'w', encoding='utf-8') as json_file:
        json.dump(data, json_file, ensure_ascii=False, default=json_encoder)


@ext_check(ext='jsonl')
def read_jsonl(file_path: str, return_iter: bool = False) -> Union[List[dict], Iterable[dict]]:
    """
    Args:
        file_path: jsonl file path
        return_iter: if the jsonl is really large, you can set `return_iter=True`
    """

    def parse_fn():
        with open(file_path, 'r', encoding='utf-8') as json_file:
            for line in json_file:
                yield json.loads(line)

    if return_iter:
        return parse_fn()
    else:
        return list(parse_fn())


@ext_check(ext='jsonl')
def dump_jsonl(data: List[Any], file_path: str, json_encoder=None):
    """
    Args:
        file_path: dump jsonl file path
        data: data to dump
        json_encoder: when you want to save a custom class, you need to set `json_encoder`
    """
    os.makedirs(os.path.dirname(file_path) or './', exist_ok=True)
    with open(file_path, 'w', encoding='utf-8') as json_file:
        for item in data:
            line = json.dumps(item, ensure_ascii=False, default=json_encoder)
            json_file.write(line + '\n')


@ext_check(ext='txt')
def read_txt(file_path: str, return_iter: bool = False) -> Union[List[str], Iterable[str]]:
    """
    Args:
        file_path: txt file path
        return_iter: if the txt file is very large, you can set return_iter=True
    """

    def parse_fn():
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                yield line.strip()

    if return_iter:
        return parse_fn()
    else:
        return list(parse_fn())


@ext_check(ext='txt')
def dump_txt(data: List[str], file_path: str):
    """
    Args:
        data: need to save
        file_path: save file path
    """
    os.makedirs(os.path.dirname(file_path) or './', exist_ok=True)
    with open(file_path, 'w', encoding='utf-8') as txt_file:
        for line in data:
            txt_file.write(line.strip() + '\n')


@ext_check(ext=['yaml', 'yml'])
def read_config(file_path: str, return_dict: bool = False) -> Union[DictConfig, dict]:
    """
    Args:
        file_path: yaml or yml config path
        return_dict: you can set `return_dict=True` to return dict, not DictConfig
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        config = OmegaConf.load(f)
    if return_dict:
        config = OmegaConf.to_container(config, resolve=True, enum_to_str=True)
    return config


@ext_check(ext=['yaml', 'yml'])
def dump_config(data: Union[dict, DictConfig], file_path: str):
    """
    Args:
        data: the config you want to save
        file_path: save path
    """
    os.makedirs(os.path.dirname(file_path) or './', exist_ok=True)
    with open(file_path, 'w', encoding='utf-8') as f:
        OmegaConf.save(data, f)


@ext_check(ext='pkl')
def read_pkl(file_path: str):
    with open(file_path, 'rb') as f:
        return pickle.load(f)


@ext_check(ext='pkl')
def dump_pkl(data: Any, file_path: str):
    os.makedirs(os.path.dirname(file_path) or './', exist_ok=True)
    with open(file_path, 'wb') as f:
        pickle.dump(data, f)


@ext_check(ext=['xlsx', 'xls'])
def read_large_excel(file_path: str) -> Iterable[dict]:
    wb = load_workbook(file_path, read_only=True)
    for sheet in wb.sheetnames[:1]:
        ws = wb[sheet]
        _iter = ws.iter_rows()
        header = [cell.value for cell in next(_iter)]
        for row in _iter:
            data = dict(zip(header, (cell.value for cell in row)))
            yield data


@ext_check(ext=['xlsx', 'xls'])
def read_excel(file_path: str, return_iter: bool = False, return_dict: bool = True) -> Union[
    pd.DataFrame, Iterable[dict]]:
    """
    Args:
        file_path:
        return_iter: if the Excel file very large, you can set `return_iter=True`
        return_dict:
    """
    if return_iter:
        return_dict = True

    if return_iter:
        return read_large_excel(file_path)
    else:
        df = pd.read_excel(file_path)
        if return_dict:
            return df.to_dict('records')
        return df


@ext_check(ext=['xlsx', 'xls'])
def dump_excel(data: pd.DataFrame, file_path: str, index: bool = True):
    """
    Args:
        data: need to save
        file_path: file save path
        index: used by `dataframe.to_excel()`
    """
    data.to_excel(file_path, index=index)


@ext_check(ext=['tsv', 'csv'])
def read_csv(
    file_path: str,
    delimiter: str = ',',
    return_iter: bool = False,
    return_dict: bool = False,
    has_header: bool = True,
) -> Union[pd.DataFrame, List[dict], Iterable[List[str]], Iterable[List[dict]]]:
    """
    Args:
        file_path: csv or tsv file path
        delimiter: if tsv file, delimiter will be changed to \t
        return_iter: if the file is large, you can set `return_iter=True`
        return_dict: if `return_dict=True`, df.to_dict('records')
        has_header: Does the csv/tsv file has header row
    """
    is_tsv = file_path.endswith('.tsv')

    if is_tsv:
        delimiter = '\t'

    def parse_fn():
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=delimiter)
            column_names = None
            if has_header and return_dict:
                column_names = [s.strip() for s in next(reader)]
            elif return_dict:
                # Dummy column names if no header but return dict
                first_row = next(reader)
                column_names = list(range(len(first_row)))
                yield dict(zip(column_names, first_row))

            for row in reader:
                if not return_dict:
                    yield row
                else:
                    if column_names is None:
                        column_names = list(range(len(row)))
                    yield dict(zip(column_names, row))

    if return_iter:
        return parse_fn()
    else:
        df = pd.read_csv(file_path, delimiter=delimiter, on_bad_lines='skip', encoding='utf-8')
        if return_dict:
            return df.to_dict('records')
        return df


@ext_check(ext=['tsv', 'csv'])
def dump_csv(data: pd.DataFrame, file_path: str, delimiter: str = ',', index: bool = True):
    """
    Args:
        data: need to save
        file_path: save path
        delimiter: if tsv file, delimiter will be changed to \t
        index: save index column or not
    """
    if file_path.endswith('.tsv'):
        delimiter = '\t'
    os.makedirs(os.path.dirname(file_path) or './', exist_ok=True)
    data.to_csv(file_path, sep=delimiter, index=index)


@ext_check(ext=['npy', 'npz'])
def read_npyz(file_path: str):
    return np.load(file_path, allow_pickle=True)


@ext_check(ext=['npy', 'npz'])
def dump_npyz(data: Any, file_path: str):
    os.makedirs(os.path.dirname(file_path) or './', exist_ok=True)
    ext = get_file_name_and_ext(file_path, with_dot=False)[1]
    if ext == 'npz':
        np.savez(file_path, data)
    else:
        np.save(file_path, data)


class FileReader:
    _FILE_EXT_TO_FUNC = {
        'json': read_json,
        'jsonl': read_jsonl,
        'txt': read_txt,
        'yaml': read_config,
        'yml': read_config,
        'pkl': read_pkl,
        'xlsx': read_excel,
        'xls': read_excel,
        'csv': read_csv,
        'tsv': read_csv,
        'npy': read_npyz,
        'npz': read_npyz,
    }

    _EXT_PARAM_MAP = {
        'json': [],
        'jsonl': ['return_iter'],
        'txt': ['return_iter'],
        'yaml': ['return_dict'],
        'yml': ['return_dict'],
        'pkl': [],
        'xlsx': ['return_iter', 'return_dict'],
        'xls': ['return_iter', 'return_dict'],
        'csv': ['return_iter', 'delimiter', 'return_dict', 'has_header'],
        'tsv': ['return_iter', 'delimiter', 'return_dict', 'has_header'],
        'npy': [],
        'npz': [],
    }

    _return_iter = False
    _return_dict = False
    _delimiter = ','
    _has_header = True

    @classmethod
    def read(
        cls,
        file_path: str,
        return_iter: bool = None,
        return_dict: bool = None,
        delimiter: str = None,
        has_header: bool = None
    ):
        if not os.path.exists(file_path) or not os.path.isfile(file_path):
            raise FileNotFoundError(f'{file_path} is not a valid file path')

        if return_iter is None:
            return_iter = cls._return_iter
        if return_dict is None:
            return_dict = cls._return_dict
        if delimiter is None:
            delimiter = cls._delimiter
        if has_header is None:
            has_header = cls._has_header

        file_ext = get_file_name_and_ext(file_path, False)[-1]
        if file_ext not in cls._FILE_EXT_TO_FUNC:
            raise ValueError(f'Unsupported file extension: {file_ext}')

        kwargs = {'file_path': file_path}

        params_needed = cls._EXT_PARAM_MAP.get(file_ext, [])
        if 'return_iter' in params_needed:
            kwargs['return_iter'] = return_iter
        if 'return_dict' in params_needed:
            kwargs['return_dict'] = return_dict
        if 'delimiter' in params_needed:
            kwargs['delimiter'] = delimiter
        if 'has_header' in params_needed:
            kwargs['has_header'] = has_header

        func = cls._FILE_EXT_TO_FUNC[file_ext]
        try:
            return func(**kwargs)
        except Exception as e:
            raise FileReadError(f'(Function {func.__name__}) Failed to read {file_path}: {e}')


class FileWriter:
    _FILE_EXT_TO_FUNC = {
        'json': dump_json,
        'jsonl': dump_jsonl,
        'txt': dump_txt,
        'yaml': dump_config,
        'yml': dump_config,
        'pkl': dump_pkl,
        'xlsx': dump_excel,
        'xls': dump_excel,
        'csv': dump_csv,
        'tsv': dump_csv,
        'npy': dump_npyz,
        'npz': dump_npyz,
    }

    _EXT_PARAM_MAP = {
        'json': ['json_encoder'],
        'jsonl': ['json_encoder'],
        'txt': [],
        'yaml': [],
        'yml': [],
        'pkl': [],
        'xlsx': ['index'],
        'xls': ['index'],
        'csv': ['index', 'delimiter'],
        'tsv': ['index', 'delimiter'],
        'npy': [],
        'npz': [],
    }

    _json_encoder = None
    _index = False
    _delimiter = ','

    @classmethod
    def write(
        cls,
        data: Any,
        file_path: str,
        json_encoder: Callable = None,
        index: bool = None,
        delimiter: str = None
    ):
        if not isinstance(file_path, str) or not file_path.strip():
            raise ValueError("Invalid file path provided.")

        if index is None:
            index = cls._index
        if delimiter is None:
            delimiter = cls._delimiter

        file_ext = get_file_name_and_ext(file_path, False)[-1]
        if file_ext not in cls._FILE_EXT_TO_FUNC:
            raise NotImplementedError(f'Not implemented for file extension: {file_ext}')

        func = cls._FILE_EXT_TO_FUNC[file_ext]
        kwargs = {
            'file_path': file_path,
            'data': data
        }

        params_needed = cls._EXT_PARAM_MAP.get(file_ext, [])
        if 'json_encoder' in params_needed:
            kwargs['json_encoder'] = json_encoder
        if 'index' in params_needed:
            kwargs['index'] = index
        if 'delimiter' in params_needed:
            kwargs['delimiter'] = delimiter

        try:
            func(**kwargs)
        except Exception as e:
            raise FileWriteError(f'(Function {func.__name__}) Failed to write {file_path}: {e}')
